__author__ = 'i20764'
import os,shutil,fileinput
import config.config
import dbconnect.dbconnection
from openpyxl import load_workbook
import pandas as pd
import csv
import glob
import re


db = dbconnect.dbconnection.dbconnect()

def getMantisDetail():
    cur = db.cursor()
    getDet = cur.execute('SELECT a.id,(select note from mantis.mantis_bugnote_text_table where id in (select max(id) from mantis.mantis_bugnote_table where bug_id = a.id )) as mantis_note from mantis.mantis_bug_table a where a.handler_id in ( ''204 '', ''330 '', ''366 '', ''374 '', ''402 '') and a.status= ''50 '' and a.project_id =  ''8''')
    getNote = cur.fetchall()
    cur.close()
    return getNote

def createdir(mantisid):
    try:
        if not os._exists(mantisid):
            os.mkdir(config.config.path+mantisid)
            print(config.config.path+mantisid)
            print('directory: '+mantisid+' created')
            wd = config.config.path+mantisid
    except:
        print('Folder already exists')
        wd=1

    return wd

def write_file(data,filename):
    with open(filename,'wb') as f:
        f.write(data)

def copy_file(src, dst, symlinks=False, ignore=None):
    for item in os.listdir(src):
        s = os.path.join(src, item)
        d = os.path.join(dst, item)
        if os.path.isdir(s):
            shutil.copytree(s, d, symlinks, ignore)
        else:
            shutil.copy2(s, d)

def get_dir():
    dir_list = os.listdir(config.config.template_path)
    return dir_list


def copyData(filepath,repofile):
    # Load in the workbook
    wb = load_workbook(filepath)

    # Get currently active sheet
    sheet = wb.active

    # get max row count
    max_row = sheet.max_row

    # get max column count
    max_column = sheet.max_column

    # iterate over all cells
    # iterate over all rows
    a = ''
    for i in range(2, max_row + 1):
        # iterate over all columns
        for j in range(1, max_column + 1):
            cell_obj = sheet.cell(row=i, column=j)
            a = a + str(cell_obj.value) + ','
        a = a + '\n'
        for fname in os.listdir(repofile):
            if fname.endswith(('.ctl','CTL','.Ctl')):
                with open(repofile+fname,"a") as myfile:
                    myfile.write(a)
                a = ''

def copy_template():
    note = getMantisDetail()
    x=1
    while x==1:
        for summary in note:
            result = summary[1].split('.')
            if len(result) <= 4:
                table_name = result[1]
                operation = result[2].split()
                id = str(summary[0])
                if ('INSERT' or 'Insert' or 'insert') in operation:
                    ind_operation = 'Insert'
                else:
                    ind_operation = 'Update'

                wd = createdir(id)

                if wd==1:
                    x=note
                    continue
                print(ind_operation)

                try:
                    cur = db.cursor()
                    cur.execute('select content from mantis.mantis_bug_file_table where bug_id = '+id+' and id in (select max(id) from mantis.mantis_bug_file_table where bug_id = '+id+')')
                    res = cur.fetchone()[0]
                    write_file(res,config.config.repo_excel+'\\'+id+'.xlsx')
                except Exception as e:
                    print(str(e))
                finally:
                    cur.close()
            else:
                continue

            copy_file(config.config.create_sqlplus,config.config.path+id, symlinks=False, ignore=None)


            list_of_dir = get_dir()

            if table_name in list_of_dir:
                listOfFile = os.listdir(config.config.template_path+table_name+'\\'+ind_operation+'\\')

                for file in listOfFile:
                    with open(config.config.template_path+table_name+'\\'+ind_operation+'\\'+file,'r') as f:
                        filedata = f.read()


                    file = re.sub('xxxxx',id,file,flags=re.IGNORECASE)
                    filedata = re.sub('xxxxx',id,filedata,flags=re.IGNORECASE)
                    wf = open(config.config.path+id+'\\'+file,'w')
                    wf.write(filedata)
                    wf.close()


                #TODO
                try:
                    file_xlsx = id+'.xlsx'
                    print(file_xlsx)
                    if os.path.isfile(config.config.repo_excel+file_xlsx):
                        copyData(config.config.repo_excel+'\\'+id+'.xlsx',config.config.path+id+'\\')
                except:
                    print('***Please confirm spreadsheets are attached or not in mantis -'+id+' and also look into mantis description.***')

            else:
                print('*****Please check note of manits - '+id+'*****')
                continue




        x = note














