__author__ = 'i20764'

import os,shutil,fileinput
import config.config
import dbconnect.dbconnection
from openpyxl import load_workbook
import pandas as pd
import csv
import glob
import re

# This function : copy the all file in one dir to another
# But we used only copy the create .bat file to respective mantis's directories
def copy_file(src, dst, symlinks=False, ignore=None):
    for item in os.listdir(src):
        s = os.path.join(src, item)
        d = os.path.join(dst, item)
        if os.path.isdir(s):
            shutil.copytree(s, d, symlinks, ignore)
        else:
            shutil.copy2(s, d)


# This function write the contents(spreadsheets)  which are attached at every mantis tickets.
# spreadsheet write to repo_excel folder with mantis_id.xlsx
def write_file(data,filename):
    with open(filename,'wb') as f:
        f.write(data)



# copyData : data copy to .ctl file of respective mantis
def copyData(filepath,repofile):

    # Load in the workbook
    wb = load_workbook(filepath)

    sheet = wb.sheetnames

    code_rule = ['Final Codes','Final Rules','FINAL CODES','FINAL RULES','FINAL CODE','Final NCS Codes','Final NCS Rules','FINAL RULE','Final Code','Final Rule','Final NCS_Codes','FINAL NCS_Codes','FINAL NCS_Rules','Final NCS_Rules','Final_NCS_Codes','Final_NCS_Rules','FINAL_NCS_CODES','FINAL_NCS_RULE','Final_NCS_Code','Final_NCS_Rule','Final NCS_CODES','Final NCS_RULES','Final','final','FINAL']

    result_sheet = list(set(sheet)&set(code_rule))
    print(result_sheet)

    if len(result_sheet) == 1:
        final_sheet = ''.join(result_sheet)
        sheet = wb[final_sheet]

        # get max row count
        max_row = sheet.max_row



        # get max column count
        max_column = sheet.max_column

        # iterate over all cells
        # iterate over all rows
        a = ''
        for i in range(2, max_row + 1):
            # iterate over all columns
            for j in range(1, max_column + 1):
                cell_obj = sheet.cell(row=i, column=j)
                result = str(cell_obj.value)
                if ',' in result:
                    result = result
                    if '"' in result:
                        result = result.replace('"', '""')
                        a = a + '"' + result + '"' + ','
                        continue
                    else:
                        a = a + '"' + result + '"' + ','
                        continue
                else:
                    a = a + str(cell_obj.value) + ','
                    continue
            a = a + '\n'
            a = a.replace('None','')
            for fname in os.listdir(repofile):
                if fname.endswith(('.ctl','CTL','.Ctl')):
                    with open(repofile+fname,"a") as myfile:
                        myfile.write(a)
                    a = ''
    else:
        for x in result_sheet:
            print("x: ", x)
            if x.lower() in ['final codes' ,'final_codes', 'final code' ,'final_ncs_codes', 'final ncs_codes', 'final ncs codes']:
                print("1st print ", x)
                # sheet = wb.active
                # print("1st print"+x)
                sheet = wb[x]

                # get max row count
                max_row = sheet.max_row

                # print(wb.sheetnames[-2:])

                # get max column count
                max_column = sheet.max_column

                # iterate over all cells
                # iterate over all rows
                a = ''
                for i in range(2, max_row + 1):
                    # iterate over all columns
                    for j in range(1, max_column + 1):
                        cell_obj = sheet.cell(row=i, column=j)
                        result = str(cell_obj.value)
                        if ',' in result:
                            result = result
                            if '"' in result:
                                result = result.replace('"', '""')
                                a = a + '"' + result + '"' + ','
                                continue
                            else:
                                a = a + '"' + result + '"' + ','
                                continue
                        else:
                            a = a + str(cell_obj.value) + ','
                            continue
                    a = a + '\n'
                    a = a.replace('None','')
                    for fname in os.listdir(repofile):
                        if fname.endswith(('CODES.ctl','codes.CTL','CODE.Ctl','InsCODES.ctl','InsCODES.CTL','InsCodes.ctl','InsCODES.ctl')):
                            with open(repofile+fname,"a") as myfile:
                                myfile.write(a)
                            a = ''

            # if result_sheet[x] == 'Final Rules' or result_sheet[x] == 'Final NCS_Rules' or result_sheet[x] == 'Final_NCS_Rules' or result_sheet[x]== 'FINAL NCS_Rules' or result_sheet[x] == 'FINAL RULES' or result_sheet[x] == 'FINAL CODE':
            else:
                print("2nd print ", x)
                # sheet = wb.active
                # print("1st print"+x)

                sheet = wb[x]

                # get max row count
                max_row = sheet.max_row

                # print(wb.sheetnames[-2:])

                # get max column count
                max_column = sheet.max_column

                # iterate over all cells
                # iterate over all rows
                a = ''
                for i in range(2, max_row + 1):
                    # iterate over all columns
                    for j in range(1, max_column + 1):
                        cell_obj = sheet.cell(row=i, column=j)
                        result = str(cell_obj.value)
                        if ',' in result:
                            result = result
                            if '"' in result:
                                result = result.replace('"', '""')
                                a = a + '"' + result + '"' + ','
                                continue
                            else:
                                a = a + '"' + result + '"' + ','
                                continue
                        else:
                            a = a + str(cell_obj.value) + ','
                            continue
                    a = a + '\n'
                    a = a.replace('None','')
                    for fname in os.listdir(repofile):
                        if fname.endswith(('RULE.ctl','RULE.CTL','RULE.Ctl','rule.ctl','Rule.ctl')):
                            with open(repofile+fname,"a") as myfile:
                                myfile.write(a)
                            a = ''