__author__ = 'i20764'
from openpyxl import load_workbook
import config.config
import dbconnect.dbconnection
import os
import re
import fnmatch

def get_header(excelfilepath):
    wb = load_workbook(excelfilepath)

    sheet = wb.sheetnames


    code_rule = ['Final Codes','Final Rules','FINAL CODES','FINAL RULES','FINAL CODE','Final NCS Codes','Final NCS Codes ','Final NCS Rules','Final NCS Rules ','FINAL RULE','Final Code','Final Rule','Final NCS_Codes','FINAL NCS_Codes','FINAL NCS_Rules','Final NCS_Rules','Final_NCS_Codes','Final_NCS_Rules','FINAL_NCS_CODES','FINAL_NCS_RULE','Final_NCS_Code','Final_NCS_Rule','Final NCS_CODES','Final NCS_RULES','Final','final','FINAL']

    result_sheet = list(set(sheet)&set(code_rule))
    print(result_sheet)
    if len(result_sheet) ==1:

        final_sheet = ''.join(result_sheet)
        sheet = wb[final_sheet]
        # get max column count
        max_column = sheet.max_column

        list_of_headers = list()

        for j in range(1, max_column + 1):
            cell_obj = sheet.cell(row=1, column=j)
            list_of_headers.append(cell_obj.value)
        final_headers = [x.upper() for x in list_of_headers]
        return final_headers
    else:
        final_headers_code = ''
        final_headers_rule = ''
        for x in result_sheet:
            if x.lower() in ['final codes' ,'final_codes', 'final code' ,'final_ncs_codes', 'final ncs_codes', 'final ncs codes']:

                sheet = wb[x]

                max_column = sheet.max_column

                list_of_headers = list()

                for j in range(1, max_column+1):
                    cell_obj = sheet.cell(row=1, column=j)
                    list_of_headers.append(cell_obj.value)
                final_headers_code = [x.upper() for x in list_of_headers]


            # if result_sheet[x] == 'Final Rules' or result_sheet[x] == 'Final NCS_Rules' or result_sheet[x] == 'Final_NCS_Rules' or result_sheet[x]== 'FINAL NCS_Rules' or result_sheet[x] == 'FINAL RULES' or result_sheet[x] == 'FINAL CODE':
            else:

                sheet = wb[x]

                max_column = sheet.max_column

                list_of_headers = list()

                for j in range(1, max_column + 1):
                    cell_obj = sheet.cell(row=1, column=j)
                    list_of_headers.append(cell_obj.value)
                final_headers_rule = [x.upper() for x in list_of_headers]

        return final_headers_code ,final_headers_rule
def get_file_count(repofilepath):
    file_count = len(fnmatch.filter(os.listdir(repofilepath),'*.[ctl][CTL][Ctl]'))
    return file_count


def get_column(repofilepath):
    file_count = get_file_count(repofilepath)
    if file_count == 1:
        data = ''
        for fname in os.listdir(repofilepath):
                    if fname.endswith(('.ctl','.CTL','.Ctl')):
                        with open(repofilepath+fname,"r") as myfile:
                            next(myfile)
                            data = myfile.read()
        single_table_columns = data.replace(',','')
        single_table_columns_next = single_table_columns.split()


        final_single_columns = [x.upper() for x in single_table_columns_next]

        return final_single_columns



    else:
        data = ''
        fields_final_codes = ''
        fields_final_rules = ''
        for fname in os.listdir(repofilepath):


                    if fname.endswith(('CODES.ctl','codes.CTL','CODE.Ctl','InsCODES.ctl','InsCODES.CTL','InsCodes.ctl','InsCODES.ctl')):
                        with open(repofilepath+fname,"r") as myfile:
                            next(myfile)
                            data_code = myfile.read()
                            columns_code = data_code.replace(',','')
                            columns_code_next = columns_code.split()
                            fields_final_codes = [x.upper() for x in columns_code_next]
                    if fname.endswith(('RULE.ctl','RULE.CTL','RULE.Ctl','rule.ctl','Rule.ctl')):
                        with open(repofilepath+fname,"r") as myfile:
                            next(myfile)
                            data_rule = myfile.read()
                            columns_rule = data_rule.replace(',','')
                            columns_rule_next = columns_rule.split()
                            fields_final_rules = [x.upper() for x in columns_rule_next]




        return fields_final_codes,fields_final_rules

def check_valid(excelfilepath,repofilepath):
    file_count = get_file_count(repofilepath)
    if file_count == 1:
        fields_single = get_column(repofilepath)
        headers_single = get_header(excelfilepath)
        match_single = list(set(headers_single)&set(fields_single))
        header_length = len(headers_single)
        match_single_length = len(match_single)
        if header_length == match_single_length:
            return 'Pass'
        else:
            return 'Columns Mismatch'
    else:
        fields_code,fields_rule = get_column(repofilepath)
        headers_code,headers_rule = get_header(excelfilepath)
        match_code = list(set(headers_code)&set(fields_code))
        match_rule = list(set(headers_rule)&set(fields_rule))
        header_code_length = len(headers_code)
        header_rule_length = len(headers_rule)
        match_code_length = len(match_code)
        match_rule_length = len(match_rule)
        if header_code_length == match_code_length and header_rule_length == match_rule_length:
            return 'Pass'
        else:
            return 'Column Mismatch'




